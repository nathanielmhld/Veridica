# -*- coding: utf-8 -*-
"""
Created on Dec 27 2016 by Nathaniel Mahowald

Contributors: Johnny Wong, Jacky Zhu
"""
# This file contains:
#   1. the Co class which stores all the info we require for each company
#   2. excelToCo function which outputs a list of Co objects corresponding to the companies provided in the Excel file
#   3. coUpdateHTML function which takes in a coList and loads the URL in the website
#       attribute of each Co object and stores the HTML content in the content attribute of the Co object
#   4. coUpdateSelenium function which attempts to use selenium to load sites that might have blocked our scraper

"""Imports"""
import openpyxl #library for pulling data from excel files
import ssl #deals with SSL Certificate Errors
import os
import pandas as pd
import numpy as np
#from urllib.request import Request, urlopen # Modules used to access websites with their URLs
#from urllib.error import URLError, HTTPError # Modules used to deal with errors with accessing websites
from http.client import IncompleteRead, BadStatusLine # catches errors that do not have a specific status code
from ssl import CertificateError # Strange error catch and bypass
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from datetime import datetime
import sys
from .utils import loadURL
from .domains import getDomain

class Co:
    """Short for COmpany, this class will contain all the data we need on each company"""
    def __init__(self, name = None, website = None, secondary = None, functional = None):
        self.name = name
        self.language = None
        self.errors = []  # List of all errors encountered during attempts
                          # on this page
        """Cleans up some of the website input"""
        if website[0:4] != 'http' and website[0:3] != 'See':
            # appends http heading if not already present and
            # checks if the provider is a subordinate of another provider
            website = 'http://' + website
        elif website[0:3] == 'See' or website[0:9] == 'Bought by':
            self.errors.append("CloudShare reference error")
        self.functional = functional # functional market of co
        self.secondary = secondary # secondary market of co
        self.website = website # URL of website
        self.keywords = {} #{string keys, string urls}
        self.oldestAccess = None
        self.recentAccess = None
        self.linkToText = {} # {link:text}
        self.linkToContent = {} # {link:content}
        self.userAccess = None

"""imports a correctly formatted spreadhseet as a list of Co objects"""
"""Depricated"""
def excelToCo(savepath, excel_file_name = 'cloudshare.xlsx', outpath = None):
    coList = []
    spreadsheet = openpyxl.load_workbook(savepath + os.sep + excel_file_name) # pulls data directly from excel file
    data = spreadsheet.get_sheet_by_name('Data') # specifies which sheet to pull data from

    for i in range(2, data.max_row):
        # data.max_row sometimes returns more than actual number of rows, so the for loop
        # iterates through empty rows, this if statement accounts for that
        if data.cell(row = i, column = 3).value is not None:
            name = data.cell(row = i, column = 2).value
            url = data.cell(row = i, column = 3).value
            secondary = data.cell(row = i, column = 18).value
            coList.append(Co(name = name, website = url, secondary = secondary))

    np.save(os.getcwd() + os.sep + 'data' + os.sep + 'data', coList)
    return coList

"""takes a list of Co objects and attempts to fill out their website content,
   errors encounters, and dates of access using basic html methods
   Ignores duplicates/references"""
def coUpdateHTML(coList, outpath = None):
    retCoList = []
    for co in coList: # iterates through every Co object
        retval = loadURL(co.website)
        if retval != None:
            co.linkToContent[co.website] = retval
            co.userAccess = time.strftime("%Y-%m-%d")
            retCoList.append(co)
            try:
                print("Obtained HTML from main site of ", co.name)
            except:
                co.name = getDomain(co.website)
                print("Obtained HTML from main site of ", co.name)
    return retCoList


"""Depricated"""
"""takes a list of Co objects and attempts to fill out the website content,
   of the cos that encountered errors before using selenium web driver
   Ignores duplicates/references"""
def coUpdateSelenium(colist, lim = -1, outpath = None):
    count = 0
    colist = colist[0:lim]
    driver = webdriver.Firefox()
    for co in colist: # iterates through every Co object that errored
        if co.errors != []:

            driver.get(co.website)
            if driver.page_source:
                count = count + 1
                co.content = driver.page_source
                print (co.name + ' is working fine with selenium')
            else:
                print ("Selenium issue")
                co.errors.append("Selenium issue")

    driver.close()
    print(str(count))
    curr_dir = os.getcwd() # gets the current directory
    np.save(curr_dir + os.sep + 'data' + os.sep + (outpath if outpath else time.strftime("%d_%m_%Y")), colist)
                                            # Updates the saved binaries
                                            # if no name is provided it uses the date
    return colist

# takes in a dictionary of urls {URL: search_term} and returns a potential Co List
def urlsToCo(url_dict):
    potentialCoList = []
    names = []
    for url in url_dict.keys(): # url_dict has the form {link: [search_term]}
        name = getDomain(url)
        if name not in names:
            names.append(name)
            new_Co = Co(name, url)
            potentialCoList.append(new_Co)
    return potentialCoList


